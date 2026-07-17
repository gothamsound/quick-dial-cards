#!/usr/bin/env python3
"""Generate quick-dial-cards-editable.xlsx from print.html (canonical source)."""
import re, html, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent
src = (ROOT/'print.html').read_text()

def txt(h):
    h = re.sub(r'<kbd>(.*?)</kbd>', r'[\1]', h)
    h = re.sub(r'<[^>]+>', '', h)
    h = html.unescape(h)
    return re.sub(r'\s+', ' ', h.replace('→','->')).strip()

cards=[]
for sheet in src.split('<section class="sheet">')[1:]:
    h1 = txt(re.search(r'<h1>(.*?)</h1>', sheet, re.S).group(1))
    if h1.startswith('Wireless Quick-Dial'): continue
    g = lambda pat: (lambda m: txt(m.group(1)) if m else '')(re.search(pat, sheet, re.S))
    def steps(cls):
        m = re.search(r'<div class="proc '+cls+r'">.*?<span class="proc-k">(.*?)</span>\s*<span class="proc-m">(.*?)</span>.*?<ol class="steps">(.*?)</ol>', sheet, re.S)
        if not m: return '','',[]
        label = txt(m.group(1)); model = txt(m.group(2))
        st = [txt(x) for x in re.findall(r'<span class="txt">(.*?)</span></li>', m.group(3), re.S)]
        return label, model, st
    txk, txm, txsteps = steps('tx'); rxk, rxm, rxsteps = steps('rx')
    cards.append(dict(
        title=h1, sub=g(r'<div class="sub">(.*?)</div>'),
        badge=g(r'<span class="tag [a-z- ]+">(.*?)</span>'),
        step=g(r'<span class="pill"><b>Step</b>(.*?)</span>'),
        rng=g(r'<span class="pill"><b>Range</b>(.*?)</span>'),
        match=g(r'<div class="linkbar"><span class="lk-t">MATCHING TX &amp; RX</span><span>(.*?)</span></div>'),
        txk=txk, txm=txm, txsteps=txsteps, rxk=rxk, rxm=rxm, rxsteps=rxsteps,
        rfrows=[(txt(a),txt(b)) for a,b in re.findall(r'<div class="rf-row"><span class="rf-k[^"]*">(.*?)</span><span class="rf-v">(.*?)</span></div>', sheet, re.S)],
        watch=[txt(x) for x in re.findall(r'<div class="watch">.*?<ul>(.*?)</ul>', sheet, re.S)[0].split('</li>') if txt(x)] if '<div class="watch">' in sheet else [],
        tables=[[[txt(cell) for cell in re.findall(r'<t[hd][^>]*>(.*?)</t[hd]>', row, re.S)] for row in re.findall(r'<tr[^>]*>(.*?)</tr>', tb, re.S)] for tb in re.findall(r'<table class="bandtable">(.*?)</table>', sheet, re.S)],
        sources=re.findall(r'<div class="sources">.*?</div>', sheet, re.S) and re.findall(r'<a href="([^"]+)">([^<]+)</a>', re.search(r'<div class="sources">(.*?)</div>', sheet, re.S).group(1)) or [],
    ))
assert len(cards)==25

TABS = {'Sennheiser evolution G3 / G4':'Sennheiser ew G3-G4','Sennheiser Digital 6000':'Sennheiser D6000',
 'Sennheiser 2000 Series':'Sennheiser 2000','Sennheiser 3000 / 5000 Series':'Sennheiser 3000-5000',
 'Shure ULX-D':'Shure ULX-D','Shure Axient Digital':'Shure Axient','Shure PSM 1000 (IEM / IFB)':'Shure PSM 1000',
 'Sony UWP-D':'Sony UWP-D','Sony DWX Digital':'Sony DWX','Wisycom MTP / MCR':'Wisycom',
 'Lectrosonics Digital Hybrid':'Lectrosonics','Lectrosonics L-Series wideband':'Lectrosonics L-Series',
 'Lectrosonics IFB (IFBlue)':'Lectrosonics IFB','Lectrosonics Duet (M2T / M2Ra)':'Lectrosonics Duet',
 'Comtek 216 (BST-25 / PR-216)':'Comtek 216','Sound Devices Astral (A20)':'Sound Devices Astral',
 'Sennheiser EW-DX':'Sennheiser EW-DX','Shure SLX-D':'Shure SLX-D','Shure UHF-R':'Shure UHF-R',
 'Lectrosonics D Squared (DBSM / DSQD)':'Lectrosonics D Squared','Zaxcom (ZMT4.5 / URX100)':'Zaxcom','Shure Wireless Workbench (WWB 6 & 7)':'Shure WWB','Teradek Bolt / Ranger (wireless video)':'Teradek','5 / 6 GHz Channel Chart (video & Wi-Fi)':'5-6 GHz Chart','2.4 GHz Auto-Hop Mics (DJI / Rode / Hollyland)':'2.4 GHz Auto-Hop'}

INK='FF16181D'; OLIVE='FF6A6F33'; MUT='FF4A4F57'; FILL=PatternFill('solid', fgColor='FFF6F6F1')
def style(ws,cell,sz=10,b=False,color=INK,wrap=True,fill=False):
    c=ws[cell]; c.font=Font(size=sz,bold=b,color=color)
    c.alignment=Alignment(wrap_text=wrap,vertical='top')
    if fill: c.fill=FILL
    return c

wb=Workbook(); ix=wb.active; ix.title='Index'
for k,w in dict(A=5,B=24,C=42,D=42,E=12,F=12).items(): ix.column_dimensions[k].width=w
style(ix,'A1',16,True).value='Wireless Quick-Dial Cards'
style(ix,'A2',10,False,MUT).value='Setting an assigned frequency on transmitter & portable receiver  -  one tab per TX/RX pair'
style(ix,'A4',9,False,MUT).value=('How to reach an arbitrary MHz:  FREE-TUNE = user bank / Tune / Frequency field   |   '
 'PICK FINE GROUP FIRST = select the fine/unlocked group, then dial   |   FREE-TUNE WITHIN BLOCK = any MHz inside the hardware block   |   '
 'CHANNEL CHART ONLY = fixed channels')
for col,h in zip('ABCD',['#','System (tab)','Transmitter','Portable receiver']):
    style(ix,f'{col}6',10,True,'FFFFFFFF').value=h
    ix[f'{col}6'].fill=PatternFill('solid',fgColor=INK)
for n,c in enumerate(cards,1):
    r=6+n
    style(ix,f'A{r}').value=n
    style(ix,f'B{r}',b=True).value=TABS[re.sub(r'\s+',' ',c['title'])]
    style(ix,f'C{r}').value=c['txm']; style(ix,f'D{r}').value=c['rxm']
r=8+len(cards)
style(ix,f'A{r}',10,True,OLIVE).value='Three rules that prevent most check-in failures'
for i,rule in enumerate([
 '1.  Transmitter and receiver must be in the same band / block.',
 '2.  On Sony UWP-D & Wisycom, select the fine / unlocked group before an arbitrary frequency is reachable (Sony DWX also offers direct FREQ INPUT).',
 '3.  On digital systems an IR-sync (RX->TX) copies the frequency onto the other unit; re-sync after changing encryption or band.']):
    style(ix,f'A{r+1+i}',9.5,False,MUT).value=rule

for c in cards:
    ws=wb.create_sheet(TABS[re.sub(r'\s+',' ',c['title'])])
    for k,w in dict(A=13,B=26,C=5,D=58,E=34).items(): ws.column_dimensions[k].width=w
    style(ws,'A1',16,True).value=c['title']
    style(ws,'A2',10,False,MUT).value=c['sub']
    style(ws,'A4',10,True,OLIVE).value='REFERENCE'
    for r,(k,v) in enumerate([('How to reach any MHz',c['badge']),('Tuning step',c['step']),
                              ('Range / band',c['rng']),('Matching TX <-> RX',c['match'])],5):
        style(ws,f'A{r}',b=True).value=k; style(ws,f'B{r}').value=v
        ws.merge_cells(f'B{r}:E{r}')
    r=10
    for label,model,st in [(c['txk'],c['txm'],c['txsteps']),(c['rxk'],c['rxm'],c['rxsteps'])]:
        if not st: continue
        style(ws,f'A{r}',10,True,OLIVE,fill=True).value=f'{label}  -  {model}'
        ws.merge_cells(f'A{r}:E{r}'); r+=1
        for col,h in zip(['A','B','C','D','E'],['Device','Model','Step','Instruction','Notes / Corrections']):
            style(ws,f'{col}{r}',9,True,MUT).value=h
        r+=1
        for i,stx in enumerate(st,1):
            if i==1:
                style(ws,f'A{r}',b=True).value=label.title()
                style(ws,f'B{r}').value=model
            style(ws,f'C{r}').value=i; style(ws,f'D{r}').value=stx
            r+=1
        r+=1
    for tb in c.get('tables',[]):
        for ri,row in enumerate(tb):
            for ci,cell in enumerate(row):
                st_=style(ws,f'{get_column_letter(ci+1)}{r}',9,ri==0,MUT if ri else 'FFFFFFFF')
                st_.value=cell
                if ri==0: ws[f'{get_column_letter(ci+1)}{r}'].fill=PatternFill('solid',fgColor=INK)
            r+=1
        r+=1
    style(ws,f'A{r}',10,True,OLIVE).value='RF POWER & SAFE POWER-ON'; r+=1
    for k,v in c['rfrows']:
        style(ws,f'A{r}',b=True).value=k; style(ws,f'B{r}').value=v; ws.merge_cells(f'B{r}:E{r}'); r+=1
    r+=1
    style(ws,f'A{r}',10,True,OLIVE).value='WATCH OUT'; r+=1
    for wtc in c['watch']:
        style(ws,f'A{r}').value='•  '+wtc; ws.merge_cells(f'A{r}:E{r}'); r+=1
    r+=1
    style(ws,f'A{r}',10,True,OLIVE).value='SOURCES'; r+=1
    for url,label in c['sources']:
        style(ws,f'A{r}',9,False,MUT).value=label; ws.merge_cells(f'A{r}:B{r}')
        cell=style(ws,f'C{r}',9,False,'FF4A6B86'); cell.value=url; ws[f'C{r}'].hyperlink=url
        ws.merge_cells(f'C{r}:E{r}'); r+=1

out=ROOT/'quick-dial-cards-editable.xlsx'
wb.save(out); print('wrote', out.name, f'({len(cards)} card tabs + index)')
